
from sentita import calculate_polarity
import openpyxl 
from translate import Translator


# Give the location of the file 
path = "m.xlsx"

# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column

# printing the value of first column
# Loop will print all values 
# of first column  

overall_positive = 0;
overall_negative = 0;
translator = Translator(to_lang="it")

best = -100;
worst = -100;

best_sentence = "";
worst_sentence = "";

for i in range(1, row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 2) 
    ##Traduzioni: 
    ##translated = translator.translate(cell_obj.value);
    sentences = [cell_obj.value];
    results, polarities = calculate_polarity(sentences)
    cell_obj = sheet_obj.cell(row = i, column = 3);
    cell_obj.value = polarities[0][0];
    cell_obj = sheet_obj.cell(row = i, column = 4);
    cell_obj.value = polarities[0][1];
    print(results);
    print('--------------------------------------------')
    overall_positive+= polarities[0][0];
    overall_negative+= polarities[0][1];

    if(polarities[0][0] > best):
        best = polarities[0][0];
        best_sentence = sentences[0];
    
    if(polarities[0][1] > worst):
        worst = polarities[0][1]
        worst_sentence =  sentences[0];
    

overall_pos = overall_positive
overall_neg = overall_negative
wb_obj.save('m.xlsx');
print('----------OVERALL-----');
print(overall_pos, overall_neg);
print('----------BEST--------')
print(best_sentence, 'WORST------\n', worst_sentence);

